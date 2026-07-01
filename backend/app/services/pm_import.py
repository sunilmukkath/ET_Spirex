"""Bulk import PM pipeline projects from Excel or CSV."""

from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

import pandas as pd
from openpyxl import Workbook, load_workbook
from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.db.models import Client, Project
from app.models.pm import PmImportResult, PmImportRowResult, PmProjectCreate
from app.services import pm_ops_store, pm_store
from app.services.pm_stage import normalize_stage
from app.services.project_import_config import (
    import_config_info,
    load_column_mapping,
    save_column_mapping,
    save_master_template,
)

HEADER_ALIASES: dict[str, str] = {
    "projectname": "project_name",
    "projectno": "project_code",
    "projectnumber": "project_code",
    "projectcode": "project_code",
    "project": "project_name",
    "name": "project_name",
    "title": "project_name",
    "clientname": "client_name",
    "client": "client_name",
    "surveyid": "limesurvey_survey_id",
    "limesurveyid": "limesurvey_survey_id",
    "limesurveysurveyid": "limesurvey_survey_id",
    "limeid": "limesurvey_survey_id",
    "studyid": "limesurvey_survey_id",
    "surveyname": "survey_name",
    "limesurvey": "survey_name",
    "limesurveyname": "survey_name",
    "studyname": "survey_name",
    "limesurveytitle": "survey_name",
    "surveytitle": "survey_name",
    "lsid": "limesurvey_survey_id",
    "lsstudyid": "limesurvey_survey_id",
    "projecttitle": "project_name",
    "study": "project_name",
    "studyname": "project_name",
    "clientorg": "client_name",
    "buyer": "client_name",
    "fy": "fiscal_year",
    "financialyear": "fiscal_year",
    "fiscalyear": "fiscal_year",
    "month": "billing_month",
    "billingmonth": "billing_month",
    "projectmonth": "billing_month",
    "projecttype": "project_type",
    "type": "project_type",
    "engagementtype": "engagement_type",
    "engagement": "engagement_type",
    "stage": "stage",
    "status": "stage",
    "owner": "owner_name",
    "ownername": "owner_name",
    "pm": "owner_name",
    "startdatetargetclosedate": "start_date",
    "startdate/targetclosedate": "start_date",
    "targetclosedate": "target_close_date",
    "closedate": "target_close_date",
    "duedate": "target_close_date",
    "budgetestimate": "budget_estimate",
    "budget": "budget_estimate",
    "projectvalue": "project_value_inr",
    "projectvalueinr": "project_value_inr",
    "projectvalue₹": "project_value_inr",
    "valueinr": "project_value_inr",
    "invoicevalue": "project_value_inr",
    "inrvalue": "project_value_inr",
    "notes": "status_notes",
    "statusnotes": "status_notes",
    "comments": "status_notes",
}

TEMPLATE_HEADERS = [
    "project_code",
    "fiscal_year",
    "billing_month",
    "project_name",
    "client_name",
    "limesurvey_survey_id",
    "survey_name",
    "project_type",
    "engagement_type",
    "stage",
    "owner_name",
    "start_date",
    "target_close_date",
    "budget_estimate",
    "project_value_inr",
    "status_notes",
]

TEMPLATE_EXAMPLE = [
    [
        "P2026_001",
        "FY2026 - 2027",
        "July'2026",
        "Brand Tracker Q3 2026",
        "Acme FMCG",
        "",
        "Acme Brand Tracker Wave 3",
        "quant",
        "tracking",
        "Fieldwork/Data Collection",
        "Sunil",
        "2026-07-01",
        "2026-09-30",
        "45000",
        "45000",
        "Monthly tracker — link by survey name if ID blank",
    ],
    [
        "P2026_002",
        "FY2026 - 2027",
        "Aug'2026",
        "Concept Test — Snacks",
        "Retail Co",
        "123456",
        "",
        "quant",
        "ad-hoc",
        "Analysis",
        "Ambika",
        "",
        "",
        "",
        "",
        "Use limesurvey_survey_id when you know the LimeSurvey study ID",
    ],
]


def project_import_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Projects"
    ws.append(TEMPLATE_HEADERS)
    for row in TEMPLATE_EXAMPLE:
        ws.append(row)
    ws.freeze_panes = "A2"
    for i, w in enumerate([14, 16, 14, 36, 22, 14, 28, 14, 14, 22, 14, 14, 18, 16, 16, 40], 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = w
    # Fix column widths properly
    from openpyxl.utils import get_column_letter

    widths = [14, 16, 14, 36, 22, 14, 28, 14, 14, 22, 14, 14, 18, 16, 16, 40]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    help_ws = wb.create_sheet("Help")
    help_ws["A1"] = "Column guide"
    lines = [
        "project_code — optional internal code, e.g. P2024_001",
        "fiscal_year — optional finance year, e.g. FY2024 - 2025",
        "billing_month — optional finance month",
        "project_name — required",
        "client_name — optional; created if new",
        "limesurvey_survey_id — LimeSurvey study ID (numeric); preferred when known",
        "survey_name — match LimeSurvey study title when ID is blank",
        "project_type — quant | qual | mixed (default quant)",
        "engagement_type — tracking | ad-hoc (default ad-hoc)",
        "stage — Proposal, Fieldwork/Data Collection, Analysis, etc.",
        "owner_name — Sunil, Ambika, Shilaja, Ravikumar, Venisha, or Samara",
        "start_date / target_close_date — YYYY-MM-DD",
        "budget_estimate — number",
        "project_value_inr — finance project value in INR; also used as budget_estimate if budget_estimate is blank",
        "status_notes — free text",
    ]
    for i, line in enumerate(lines, 2):
        help_ws.cell(i, 1, line)
    help_ws.column_dimensions["A"].width = 72

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").strip().lower())


def _map_headers(raw_headers: list[Any]) -> dict[int, str]:
    custom = load_column_mapping()
    mapped: dict[int, str] = {}
    for idx, raw in enumerate(raw_headers):
        raw_str = str(raw or "").strip()
        if raw_str in custom:
            mapped[idx] = custom[raw_str]
            continue
        key = HEADER_ALIASES.get(_norm_header(raw))
        if key:
            mapped[idx] = key
    return mapped


def _suggest_field_for_header(header: str) -> str | None:
    custom = load_column_mapping()
    if header in custom:
        return custom[header]
    return HEADER_ALIASES.get(_norm_header(header))


def _load_dataframe(data: bytes, filename: str) -> pd.DataFrame:
    buf = io.BytesIO(data)
    lower = (filename or "").lower()
    if lower.endswith(".csv") or (not lower.endswith((".xls", ".xlsx", ".xlsm")) and data[:2] != b"PK"):
        df = pd.read_csv(buf)
    elif lower.endswith(".xls"):
        df = pd.read_excel(buf, engine="xlrd")
    else:
        df = pd.read_excel(buf, engine="openpyxl")
    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _cell_to_str(val: Any) -> str:
    if isinstance(val, pd.Timestamp):
        return val.date().isoformat()
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date):
        return val.isoformat()
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    text = str(val).strip()
    if text.endswith(".0") and text[:-2].lstrip("-").isdigit():
        return text[:-2]
    return text


def _parse_rows_from_dataframe(df: pd.DataFrame) -> list[dict[str, str]]:
    if df.empty:
        return []
    headers = list(df.columns)
    col_map = _map_headers(headers)
    if "project_name" not in col_map.values():
        raise ValueError(
            "Sheet must include a project name column. Upload your master sheet once under "
            "Operations → Configure import, or use a project_name column."
        )

    out: list[dict[str, str]] = []
    for _, series in df.iterrows():
        if series.isna().all():
            continue
        item: dict[str, str] = {}
        for idx, field in col_map.items():
            if idx >= len(headers):
                continue
            val = series.iloc[idx]
            if pd.isna(val):
                continue
            text = _cell_to_str(val)
            if text and text.lower() != "nan":
                item[field] = text
        if item.get("project_name"):
            out.append(item)
    return out


def inspect_project_sheet(data: bytes, *, filename: str = "") -> dict[str, Any]:
    df = _load_dataframe(data, filename)
    headers = list(df.columns)
    suggested = {h: _suggest_field_for_header(h) for h in headers}
    samples: list[dict[str, str]] = []
    for _, series in df.head(5).iterrows():
        row: dict[str, str] = {}
        for h in headers:
            val = series[h]
            if pd.isna(val):
                row[h] = ""
            elif isinstance(val, pd.Timestamp):
                row[h] = val.date().isoformat()
            else:
                row[h] = str(val).strip()
        samples.append(row)
    return {
        "headers": headers,
        "suggested_column_map": suggested,
        "sample_rows": samples,
        "row_count": int(len(df)),
    }


def configure_import_from_master(data: bytes, *, filename: str = "") -> dict[str, Any]:
    inspection = inspect_project_sheet(data, filename=filename)
    column_map = {
        h: field
        for h, field in inspection["suggested_column_map"].items()
        if field
    }
    if "project_name" not in column_map.values():
        raise ValueError("Could not detect a project name column — rename a column to 'Project name' or similar.")
    save_column_mapping(column_map, source_filename=filename)
    save_master_template(data)
    return import_config_info()


def _parse_rows_from_xlsx(data: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    col_map = _map_headers(list(header_row))
    if "project_name" not in col_map.values():
        raise ValueError("Sheet must include a project_name column (or alias: project, name, title)")

    out: list[dict[str, str]] = []
    for excel_row in rows_iter:
        if not excel_row or all(v is None or str(v).strip() == "" for v in excel_row):
            continue
        item: dict[str, str] = {}
        for idx, field in col_map.items():
            val = excel_row[idx] if idx < len(excel_row) else None
            if val is None:
                continue
            if isinstance(val, datetime):
                item[field] = val.date().isoformat()
            elif isinstance(val, date):
                item[field] = val.isoformat()
            else:
                item[field] = str(val).strip()
        if item.get("project_name"):
            out.append(item)
    return out


def _parse_rows_from_csv(data: bytes) -> list[dict[str, str]]:
    text = data.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    try:
        header_row = next(reader)
    except StopIteration:
        return []
    col_map = _map_headers(header_row)
    if "project_name" not in col_map.values():
        raise ValueError("CSV must include a project_name column (or alias: project, name, title)")

    out: list[dict[str, str]] = []
    for row in reader:
        if not row or all(not str(v).strip() for v in row):
            continue
        item: dict[str, str] = {}
        for idx, field in col_map.items():
            if idx < len(row) and str(row[idx]).strip():
                item[field] = str(row[idx]).strip()
        if item.get("project_name"):
            out.append(item)
    return out


def parse_project_sheet(data: bytes, *, filename: str = "") -> list[dict[str, str]]:
    lower = (filename or "").lower()
    if lower.endswith(".xls") or lower.endswith(".xlsx") or lower.endswith(".xlsm") or lower.endswith(".csv"):
        return _parse_rows_from_dataframe(_load_dataframe(data, filename))
    if data[:2] != b"PK":
        return _parse_rows_from_csv(data)
    return _parse_rows_from_xlsx(data)


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(value.strip(), fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(value: str | None) -> Decimal | None:
    if not value:
        return None
    cleaned = re.sub(r"[£$,]", "", value.strip())
    try:
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _load_lime_surveys() -> list[dict[str, Any]]:
    try:
        from app.lime_client import list_projects

        return list_projects() or []
    except Exception:
        return []


def _survey_indexes(surveys: list[dict[str, Any]]) -> tuple[dict[int, str], dict[str, list[int]]]:
    by_id: dict[int, str] = {}
    by_title: dict[str, list[int]] = {}
    for s in surveys:
        sid = s.get("id")
        if sid is None:
            continue
        try:
            iid = int(sid)
        except (TypeError, ValueError):
            continue
        title = str(s.get("title") or s.get("name") or "").strip()
        by_id[iid] = title
        norm = title.lower()
        by_title.setdefault(norm, []).append(iid)
    return by_id, by_title


def _resolve_survey_id(
    row: dict[str, str],
    by_id: dict[int, str],
    by_title: dict[str, list[int]],
) -> tuple[int | None, str | None]:
    raw_id = row.get("limesurvey_survey_id", "").strip()
    if raw_id:
        try:
            sid = int(float(raw_id))
        except ValueError:
            return None, f"Invalid survey ID '{raw_id}'"
        if sid not in by_id:
            return None, f"LimeSurvey study {sid} not found in connected LimeSurvey"
        return sid, None

    name = row.get("survey_name", "").strip()
    if not name:
        return None, None

    norm = name.lower()
    if norm in by_title:
        matches = by_title[norm]
        if len(matches) == 1:
            return matches[0], None
        return None, f"Multiple LimeSurvey studies match '{name}' — set limesurvey_survey_id"

    partial = [sid for title, ids in by_title.items() for sid in ids if norm in title or title in norm]
    if len(partial) == 1:
        return partial[0], f"Matched survey by partial title → ID {partial[0]}"
    if len(partial) > 1:
        return None, f"Ambiguous survey name '{name}' — set limesurvey_survey_id"

    return None, f"No LimeSurvey study matched '{name}'"


def _get_or_create_client(session: Session, name: str | None) -> Client | None:
    if not name or not name.strip():
        return None
    clean = name.strip()
    existing = session.scalar(select(Client).where(Client.client_name == clean))
    if existing:
        return existing
    from app.models.pm import ClientCreate

    out = pm_ops_store.create_client(session, ClientCreate(client_name=clean))
    return session.get(Client, out.client_id)


def _find_existing_project(session: Session, row: dict[str, str]) -> Project | None:
    code = (row.get("project_code") or "").strip()
    if code:
        found = session.scalar(select(Project).where(Project.project_code == code))
        if found:
            return found
    name = (row.get("project_name") or "").strip()
    if not name:
        return None
    return session.scalar(
        select(Project).where(func.lower(Project.project_name) == name.lower())
    )


def _sync_project_from_row(
    session: Session,
    project: Project,
    row: dict[str, str],
    *,
    survey_id: int | None,
    survey_note: str | None,
) -> list[str]:
    """Update an existing project from import row. Returns change notes."""
    notes: list[str] = []

    raw_stage = row.get("stage")
    if raw_stage and str(raw_stage).strip().lower() not in ("", "nan"):
        stage = normalize_stage(raw_stage)
        if project.stage != stage:
            project.stage = stage
            notes.append(f"stage → {stage}")

    for field, attr in (
        ("fiscal_year", "fiscal_year"),
        ("billing_month", "billing_month"),
        ("project_code", "project_code"),
        ("status_notes", "status_notes"),
    ):
        val = (row.get(field) or "").strip()
        if val and getattr(project, attr) != val:
            setattr(project, attr, val)
            notes.append(attr)

    project_value_inr = _parse_decimal(row.get("project_value_inr"))
    budget_estimate = _parse_decimal(row.get("budget_estimate")) or project_value_inr
    if project_value_inr is not None and project.project_value_inr != project_value_inr:
        project.project_value_inr = project_value_inr
        notes.append("project_value_inr")
    if budget_estimate is not None and project.budget_estimate != budget_estimate:
        project.budget_estimate = budget_estimate
        notes.append("budget")

    start = _parse_date(row.get("start_date"))
    if start and project.start_date != start:
        project.start_date = start
        notes.append("start_date")
    target = _parse_date(row.get("target_close_date"))
    if target and project.target_close_date != target:
        project.target_close_date = target
        notes.append("target_close_date")

    owner_name = (row.get("owner_name") or "").strip()
    if owner_name:
        from app.services.pm_store import _resolve_owner_id

        owner_id = _resolve_owner_id(session, None, owner_name)
        if owner_id and project.owner_id != owner_id:
            project.owner_id = owner_id
            notes.append(f"owner → {owner_name}")

    if survey_id is not None and project.limesurvey_survey_id != survey_id:
        pm_ops_store.link_survey(session, project.project_id, survey_id)
        notes.append(f"linked survey #{survey_id}")
    elif survey_note:
        notes.append(survey_note)

    session.flush()
    return notes


def _survey_already_linked(session: Session, survey_id: int) -> str | None:
    for row in session.scalars(select(Project)).all():
        if survey_id in pm_ops_store.linked_survey_ids_for_project(row):
            return row.project_name
    return None


def import_projects_from_sheet(
    session: Session,
    data: bytes,
    *,
    filename: str = "",
    skip_duplicates: bool = True,
) -> PmImportResult:
    rows = parse_project_sheet(data, filename=filename)
    if not rows:
        raise ValueError("No project rows found in the file")

    surveys = _load_lime_surveys()
    by_id, by_title = _survey_indexes(surveys)

    existing_names = {
        name.lower()
        for name in session.scalars(select(Project.project_name)).all()
    }

    results: list[PmImportRowResult] = []
    created = skipped = updated = errors = 0

    for i, row in enumerate(rows, start=2):
        name = row["project_name"].strip()

        survey_id, survey_note = _resolve_survey_id(row, by_id, by_title)
        if survey_note and survey_id is None:
            note_lower = survey_note.lower()
            if "invalid survey id" in note_lower or "not found in connected" in note_lower:
                errors += 1
                results.append(
                    PmImportRowResult(
                        row_number=i,
                        project_name=name,
                        status="error",
                        message=survey_note,
                    )
                )
                continue
            if "ambiguous" in note_lower or "multiple" in note_lower:
                errors += 1
                results.append(
                    PmImportRowResult(
                        row_number=i,
                        project_name=name,
                        status="error",
                        message=survey_note,
                    )
                )
                continue

        if skip_duplicates:
            existing = _find_existing_project(session, row)
            if existing:
                if survey_id is not None:
                    linked_to = _survey_already_linked(session, survey_id)
                    if linked_to and linked_to != existing.project_name:
                        errors += 1
                        results.append(
                            PmImportRowResult(
                                row_number=i,
                                project_name=name,
                                status="error",
                                limesurvey_survey_id=survey_id,
                                message=f"Survey {survey_id} already linked to '{linked_to}'",
                            )
                        )
                        continue
                change_notes = _sync_project_from_row(
                    session, existing, row, survey_id=survey_id, survey_note=survey_note
                )
                if change_notes:
                    updated += 1
                    results.append(
                        PmImportRowResult(
                            row_number=i,
                            project_name=name,
                            status="updated",
                            project_id=existing.project_id,
                            limesurvey_survey_id=existing.limesurvey_survey_id,
                            message="Updated: " + ", ".join(change_notes),
                        )
                    )
                else:
                    skipped += 1
                    results.append(
                        PmImportRowResult(
                            row_number=i,
                            project_name=name,
                            status="skipped",
                            message="Already up to date",
                        )
                    )
                continue

        if survey_id is not None:
            linked_to = _survey_already_linked(session, survey_id)
            if linked_to:
                errors += 1
                results.append(
                    PmImportRowResult(
                        row_number=i,
                        project_name=name,
                        status="error",
                        limesurvey_survey_id=survey_id,
                        message=f"Survey {survey_id} already linked to '{linked_to}'",
                    )
                )
                continue

        project_value_inr = _parse_decimal(row.get("project_value_inr"))
        budget_estimate = _parse_decimal(row.get("budget_estimate")) or project_value_inr
        ptype = (row.get("project_type") or "quant").strip().lower()
        if ptype not in ("quant", "qual", "mixed"):
            ptype = "quant"
        etype = (row.get("engagement_type") or "ad-hoc").strip().lower()
        if etype not in ("tracking", "ad-hoc"):
            etype = "ad-hoc"
        stage = normalize_stage(row.get("stage"))

        try:
            client = _get_or_create_client(session, row.get("client_name"))
            project = pm_store.create_project(
                session,
                PmProjectCreate(
                    project_name=name,
                    client_id=client.client_id if client else None,
                    project_type=ptype,  # type: ignore[arg-type]
                    engagement_type=etype,  # type: ignore[arg-type]
                    stage=stage,  # type: ignore[arg-type]
                    owner_name=row.get("owner_name"),
                    limesurvey_survey_id=survey_id,
                    project_code=row.get("project_code"),
                    fiscal_year=row.get("fiscal_year"),
                    billing_month=row.get("billing_month"),
                    start_date=_parse_date(row.get("start_date")),
                    target_close_date=_parse_date(row.get("target_close_date")),
                    budget_estimate=budget_estimate,
                    project_value_inr=project_value_inr,
                    status_notes=row.get("status_notes"),
                ),
            )
            if survey_id is not None:
                pm_ops_store.link_survey(session, project.project_id, survey_id)

            existing_names.add(name.lower())
            created += 1
            msg_parts = ["Created"]
            if survey_id:
                msg_parts.append(f"linked to LimeSurvey #{survey_id}")
            elif survey_note:
                msg_parts.append(survey_note)
            else:
                msg_parts.append("no survey link (add survey_name or limesurvey_survey_id)")
            results.append(
                PmImportRowResult(
                    row_number=i,
                    project_name=name,
                    status="created",
                    project_id=project.project_id,
                    limesurvey_survey_id=survey_id,
                    message=" — ".join(msg_parts),
                )
            )
        except Exception as exc:
            errors += 1
            results.append(
                PmImportRowResult(
                    row_number=i,
                    project_name=name,
                    status="error",
                    message=str(exc),
                )
            )

    return PmImportResult(
        total_rows=len(rows),
        created=created,
        updated=updated,
        skipped=skipped,
        errors=errors,
        rows=results,
    )

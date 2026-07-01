from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

DATA_SHEET_NAME = "Crosstab"
INDEX_SHEET_NAME = "Index"
TABLE_GAP_ROWS = 4

# Elastic Tree brand palette (ARGB hex without #)
_NAVY = "0B2545"
_YELLOW = "F0B429"
_HEADER_BG = "E8EEF4"
_ZEBRA = "F8FAFC"
_TOTAL_BG = "FEF9E7"
_BORDER = "CBD5E1"
_MUTED = "64748B"
_LINK = "0563C1"

_THIN = Side(style="thin", color=_BORDER)
_MEDIUM = Side(style="medium", color=_NAVY)
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_MEDIUM)

_FONT_TITLE = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
_FONT_SECTION = Font(bold=True, size=11, color=_NAVY, name="Calibri")
_FONT_META = Font(italic=True, size=9, color=_MUTED, name="Calibri")
_FONT_HEADER = Font(bold=True, size=10, color=_NAVY, name="Calibri")
_FONT_BODY = Font(size=10, color="1E293B", name="Calibri")
_FONT_TOTAL = Font(bold=True, size=10, color=_NAVY, name="Calibri")
_FONT_INDEX_HDR = Font(bold=True, size=10, color="FFFFFF", name="Calibri")
_FONT_LINK = Font(size=10, color=_LINK, underline="single", name="Calibri")

_FILL_TITLE = PatternFill("solid", fgColor=_NAVY)
_FILL_HEADER = PatternFill("solid", fgColor=_HEADER_BG)
_FILL_INDEX_HDR = PatternFill("solid", fgColor=_NAVY)
_FILL_ZEBRA = PatternFill("solid", fgColor=_ZEBRA)
_FILL_TOTAL = PatternFill("solid", fgColor=_TOTAL_BG)

_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")


def banner_result_to_excel(result: dict[str, Any]) -> bytes:
    wb = Workbook()
    index_ws = wb.active
    index_ws.title = INDEX_SHEET_NAME
    data_ws = wb.create_sheet(title=DATA_SHEET_NAME)

    tables = _tables_from_result(result)
    if not tables:
        data_ws["A1"] = result.get("error", "No data")
        _write_index_header(index_ws)
        index_ws.cell(3, 1, "No tables exported")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    _write_index_header(index_ws)
    row = 1
    index_row = 3

    for i, table in enumerate(tables, start=1):
        title = _table_display_title(table, i)
        anchor_row = row

        index_ws.cell(index_row, 1, i).alignment = _ALIGN_CENTER
        index_cell = index_ws.cell(index_row, 2, title)
        index_cell.hyperlink = Hyperlink(
            ref=index_cell.coordinate,
            location=f"'{DATA_SHEET_NAME}'!A{anchor_row}",
            display=title,
        )
        index_cell.font = _FONT_LINK
        index_cell.alignment = _ALIGN_LEFT
        go_cell = index_ws.cell(index_row, 3, f"A{anchor_row}")
        go_cell.font = Font(size=9, color=_MUTED, name="Calibri")
        go_cell.alignment = _ALIGN_CENTER
        index_row += 1

        row = _write_table(data_ws, table, result, start_row=row)
        row += TABLE_GAP_ROWS

    index_ws.column_dimensions["A"].width = 5
    index_ws.column_dimensions["B"].width = 58
    index_ws.column_dimensions["C"].width = 10
    index_ws.freeze_panes = "A3"

    data_ws.sheet_view.showGridLines = False

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _tables_from_result(result: dict[str, Any]) -> list[dict[str, Any]]:
    if result.get("table_type") == "multi" and result.get("tables"):
        return [t for t in result["tables"] if not t.get("error")]
    if result.get("error"):
        return []
    return [result]


def _write_index_header(ws) -> None:
    ws.merge_cells("A1:C1")
    title = ws.cell(1, 1, "Crosstab export")
    title.font = Font(bold=True, size=14, color=_NAVY, name="Calibri")
    title.alignment = Alignment(horizontal="left", vertical="center")

    for col, label in enumerate(["#", "Question", "Go to"], start=1):
        cell = ws.cell(2, col, label)
        cell.font = _FONT_INDEX_HDR
        cell.fill = _FILL_INDEX_HDR
        cell.alignment = _ALIGN_CENTER if col != 2 else _ALIGN_LEFT
        cell.border = _CELL_BORDER


def _table_display_title(table: dict[str, Any], index: int) -> str:
    row = table.get("row_variable") or {}
    text = str(row.get("text") or table.get("row_header") or "").strip()
    code = str(row.get("code") or "").strip()
    if text and code:
        return f"{code}: {text}"[:120]
    if text:
        return text[:120]
    if code:
        return code[:120]
    return f"Table {index}"


def _style_cell(
    cell,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = None,
    number_format: str | None = None,
) -> None:
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def _write_table(
    ws,
    table: dict[str, Any],
    meta: dict[str, Any],
    *,
    start_row: int = 1,
) -> int:
    row = start_row
    row_var = table.get("row_variable") or {}
    title_text = row_var.get("text") or table.get("row_header") or "Crosstab"

    title_cell = ws.cell(row, 1, title_text)
    _style_cell(title_cell, font=_FONT_TITLE, fill=_FILL_TITLE, alignment=_ALIGN_LEFT)
    row += 1

    code = str(row_var.get("code") or "").strip()
    if code:
        meta_cell = ws.cell(row, 1, f"Variable: {code}")
        _style_cell(meta_cell, font=_FONT_META, alignment=_ALIGN_LEFT)
        row += 1

    banners = table.get("banner_variables") or []
    if banners:
        banner_text = "Banners: " + ", ".join(b.get("text", b.get("code", "")) for b in banners)
        banner_cell = ws.cell(row, 1, banner_text)
        _style_cell(banner_cell, font=_FONT_META, alignment=_ALIGN_LEFT)
        row += 1

    base_n = table.get("base_n") or meta.get("base_n")
    if base_n is not None:
        base_cell = ws.cell(row, 1, f"Base (weighted n): {base_n:,}" if isinstance(base_n, int) else f"Base: {base_n}")
        _style_cell(base_cell, font=_FONT_META, alignment=_ALIGN_LEFT)
        row += 1

    conf = meta.get("confidence_level") or table.get("confidence_level")
    if conf and meta.get("show_significance"):
        sig_cell = ws.cell(row, 1, f"Significance vs Total at {int(conf * 100)}% confidence")
        _style_cell(sig_cell, font=_FONT_META, alignment=_ALIGN_LEFT)
        row += 1

    row += 1

    if table.get("table_type") == "array" and table.get("sections"):
        for section in table["sections"]:
            row = _write_distribution_sheet_section(ws, section, meta, start_row=row)
            row += 2
        _merge_title_row(ws, start_row, row - 1)
        return row

    end_row = _write_distribution_sheet_section(ws, table, meta, start_row=row)
    _merge_title_row(ws, start_row, end_row - 1)
    return end_row


def _merge_title_row(ws, title_row: int, last_row: int) -> None:
    max_col = 1
    for r in range(title_row, min(last_row + 1, title_row + 40)):
        for c in range(1, 30):
            if ws.cell(r, c).value not in (None, ""):
                max_col = max(max_col, c)
    if max_col > 1:
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=max_col)


def _write_distribution_sheet_section(
    ws,
    table: dict[str, Any],
    meta: dict[str, Any],
    *,
    start_row: int,
) -> int:
    row = start_row
    if table.get("subquestion"):
        sub_cell = ws.cell(row, 1, table["subquestion"])
        _style_cell(sub_cell, font=_FONT_SECTION, alignment=_ALIGN_LEFT)
        row += 1

    headers = table.get("headers") or []
    show_counts = meta.get("show_counts", True)
    show_col_pct = meta.get("show_col_pct", True)
    show_row_pct = meta.get("show_row_pct", False)

    header_row = row
    col = 1
    label_cell = ws.cell(row, col, table.get("row_header") or "Answer")
    _style_cell(label_cell, font=_FONT_HEADER, fill=_FILL_HEADER, alignment=_ALIGN_LEFT, border=_HEADER_BORDER)
    col += 1

    col_plan: list[tuple[str, str]] = [("label", "")]
    for h in headers:
        label = h.get("label", "")
        if show_counts and show_col_pct:
            for suffix, kind in ((f"{label} (n)", "count"), (f"{label} (%)", "pct")):
                cell = ws.cell(row, col, suffix)
                _style_cell(cell, font=_FONT_HEADER, fill=_FILL_HEADER, alignment=_ALIGN_CENTER, border=_HEADER_BORDER)
                col_plan.append((kind, h.get("key", "")))
                col += 1
        elif show_col_pct:
            cell = ws.cell(row, col, f"{label} (%)")
            _style_cell(cell, font=_FONT_HEADER, fill=_FILL_HEADER, alignment=_ALIGN_CENTER, border=_HEADER_BORDER)
            col_plan.append(("pct", h.get("key", "")))
            col += 1
        elif show_counts:
            cell = ws.cell(row, col, f"{label} (n)")
            _style_cell(cell, font=_FONT_HEADER, fill=_FILL_HEADER, alignment=_ALIGN_CENTER, border=_HEADER_BORDER)
            col_plan.append(("count", h.get("key", "")))
            col += 1
        else:
            cell = ws.cell(row, col, label)
            _style_cell(cell, font=_FONT_HEADER, fill=_FILL_HEADER, alignment=_ALIGN_CENTER, border=_HEADER_BORDER)
            col_plan.append(("text", h.get("key", "")))
            col += 1
        if show_row_pct and h.get("key") == "total":
            cell = ws.cell(row, col, "Row %")
            _style_cell(cell, font=_FONT_HEADER, fill=_FILL_HEADER, alignment=_ALIGN_CENTER, border=_HEADER_BORDER)
            col_plan.append(("row_pct", "total"))
            col += 1

    max_col = col - 1
    row += 1
    data_start = row

    for data_idx, data_row in enumerate(table.get("rows") or []):
        is_total = bool(data_row.get("is_total"))
        stripe = _FILL_TOTAL if is_total else (_FILL_ZEBRA if data_idx % 2 == 1 else None)
        body_font = _FONT_TOTAL if is_total else _FONT_BODY
        col = 1

        label = ws.cell(row, col, data_row.get("label", ""))
        _style_cell(label, font=body_font, fill=stripe, alignment=_ALIGN_LEFT, border=_CELL_BORDER)
        col += 1

        cell_idx = 0
        for i, cell_data in enumerate(data_row.get("cells") or []):
            h = headers[i] if i < len(headers) else {}
            sig = cell_data.get("sig")

            if show_counts and show_col_pct:
                count_cell = ws.cell(row, col, cell_data.get("count", ""))
                _style_cell(
                    count_cell,
                    font=body_font,
                    fill=stripe,
                    alignment=_ALIGN_RIGHT,
                    border=_CELL_BORDER,
                    number_format="#,##0",
                )
                col += 1

                pct_val = cell_data.get("col_pct")
                pct_cell = ws.cell(row, col)
                if sig:
                    pct_cell.value = f"{pct_val}%{(' ' + sig) if sig else ''}"
                    _style_cell(pct_cell, font=body_font, fill=stripe, alignment=_ALIGN_RIGHT, border=_CELL_BORDER)
                elif pct_val is not None and pct_val != "":
                    pct_cell.value = float(pct_val) / 100.0
                    _style_cell(
                        pct_cell,
                        font=body_font,
                        fill=stripe,
                        alignment=_ALIGN_RIGHT,
                        border=_CELL_BORDER,
                        number_format="0.0%",
                    )
                col += 1
            elif show_col_pct:
                pct_val = cell_data.get("col_pct")
                pct_cell = ws.cell(row, col)
                if sig:
                    pct_cell.value = f"{pct_val}%{(' ' + sig) if sig else ''}"
                elif pct_val is not None and pct_val != "":
                    pct_cell.value = float(pct_val) / 100.0
                    _style_cell(
                        pct_cell,
                        font=body_font,
                        fill=stripe,
                        alignment=_ALIGN_RIGHT,
                        border=_CELL_BORDER,
                        number_format="0.0%",
                    )
                else:
                    _style_cell(pct_cell, font=body_font, fill=stripe, alignment=_ALIGN_RIGHT, border=_CELL_BORDER)
                col += 1
            else:
                parts = []
                if show_counts:
                    parts.append(str(cell_data.get("count", "")))
                if show_col_pct:
                    parts.append(f"{cell_data.get('col_pct', '')}%")
                if sig:
                    parts.append(sig)
                text_cell = ws.cell(row, col, " ".join(parts) if parts else "")
                _style_cell(text_cell, font=body_font, fill=stripe, alignment=_ALIGN_RIGHT, border=_CELL_BORDER)
                col += 1

            if show_row_pct and h.get("key") == "total" and cell_data.get("row_pct") is not None:
                rp = ws.cell(row, col, float(cell_data["row_pct"]) / 100.0)
                _style_cell(
                    rp,
                    font=body_font,
                    fill=stripe,
                    alignment=_ALIGN_RIGHT,
                    border=_CELL_BORDER,
                    number_format="0.0%",
                )
                col += 1

            cell_idx += 1

        row += 1

    _auto_column_widths(ws, header_row, row - 1, max_col)

    return row


def _auto_column_widths(ws, start_row: int, end_row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        letter = get_column_letter(c)
        max_len = 10 if c == 1 else 8
        for r in range(start_row, end_row + 1):
            val = ws.cell(r, c).value
            if val is None:
                continue
            max_len = max(max_len, min(len(str(val)) + 2, 48))
        ws.column_dimensions[letter].width = max_len

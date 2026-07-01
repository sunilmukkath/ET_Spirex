"""Tests for crosstab Excel export."""

import io

from openpyxl import load_workbook

from app.services.excel_export import banner_result_to_excel


def _sample_table(code: str, text: str) -> dict:
    return {
        "row_variable": {"code": code, "text": text},
        "row_header": "Answer",
        "headers": [{"label": "Total", "key": "total"}],
        "rows": [
            {
                "label": "Yes",
                "cells": [{"count": 10, "col_pct": 50.0}],
            },
            {
                "label": "No",
                "cells": [{"count": 10, "col_pct": 50.0}],
            },
        ],
    }


def test_multi_table_export_single_sheet_with_index():
    result = {
        "table_type": "multi",
        "show_counts": True,
        "show_col_pct": True,
        "tables": [
            _sample_table("Q1", "Awareness"),
            _sample_table("Q2", "Consideration"),
        ],
    }

    raw = banner_result_to_excel(result)
    wb = load_workbook(io.BytesIO(raw))

    assert wb.sheetnames == ["Index", "Crosstab"]
    assert wb["Index"]["B3"].value == "Q1: Awareness"
    assert wb["Index"]["B4"].value == "Q2: Consideration"
    assert wb["Index"]["B3"].hyperlink.location == "'Crosstab'!A1"
    assert wb["Index"]["B4"].hyperlink.location == "'Crosstab'!A11"

    data = wb["Crosstab"]
    assert data["A1"].value == "Awareness"
    assert data["A11"].value == "Consideration"
    assert data["C5"].number_format == "0.0%"


def test_single_table_export_has_index():
    table = _sample_table("Q1", "Satisfaction")

    raw = banner_result_to_excel({**table, "show_counts": True, "show_col_pct": True})
    wb = load_workbook(io.BytesIO(raw))

    assert wb.sheetnames == ["Index", "Crosstab"]
    assert wb["Index"]["B3"].value == "Q1: Satisfaction"
    assert wb["Index"]["B3"].hyperlink.location == "'Crosstab'!A1"
    assert wb["Crosstab"]["A1"].value == "Satisfaction"
